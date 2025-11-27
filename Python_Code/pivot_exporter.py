import os
import pandas as pd
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font as OpenPyXLFont, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class PivotExporter:
    """Handles exporting the pivot table to an Excel file."""
    def __init__(self, pivot_tab):
        self.pivot_tab = pivot_tab
        self.logger = pivot_tab.logger
    
    def export_pivot(self):
        """Export the pivot table to an Excel file with formatting matching the UI."""
        if self.pivot_tab.current_view_df is None or self.pivot_tab.current_view_df.empty:
            self.logger.warning("No data to export")
            QMessageBox.warning(self.pivot_tab, "Warning", "No data to export!")
            return

        try:
            # Verify required methods exist
            if not hasattr(self.pivot_tab, 'is_numeric') or not hasattr(self.pivot_tab, 'format_value'):
                self.logger.error("PivotTab is missing required methods: 'is_numeric' or 'format_value'")
                raise AttributeError("PivotTab is missing required methods: 'is_numeric' or 'format_value'")

            # Get save file path
            file_path = QFileDialog.getSaveFileName(self.pivot_tab, "Save Pivot Table", "pivot_table.xlsx", "Excel files (*.xlsx)")[0]
            if not file_path:
                self.logger.debug("Export cancelled by user")
                self.pivot_tab.status_label.setText("Export cancelled")
                return

            # Prepare data for export
            df = self.pivot_tab.current_view_df.copy()
            export_rows = [row for _, row in df.iterrows()]
            export_index = list(df.index)

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Pivot Table"

            # Define styles
            header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            first_col_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
            odd_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            even_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            header_font = OpenPyXLFont(name="Segoe UI", size=12, bold=True)
            cell_font = OpenPyXLFont(name="Segoe UI", size=12)
            cell_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

            # Write headers
            headers = list(df.columns)
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = cell_align
                c.border = thin_border
                ws.column_dimensions[get_column_letter(ci)].width = 15

            # Write data
            for row_idx, row in enumerate(export_rows, start=2):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=ci)
                    if pd.isna(val):
                        cell.value = None  # Store NaN as empty cell
                    else:
                        try:
                            # Try to convert to numeric
                            numeric_value = float(val)
                            # Calculate number of decimal places
                            str_val = str(val).rstrip('0').rstrip('.')
                            decimal_places = len(str_val.split('.')[-1]) if '.' in str_val else 0
                            cell.value = numeric_value  # Store as numeric
                            cell.number_format = f"0.{'0' * decimal_places}" if decimal_places > 0 else "0"
                        except (ValueError, TypeError):
                            # Store non-numeric values as strings
                            cell.value = str(val)
                    cell.font = cell_font
                    cell.alignment = cell_align
                    cell.border = thin_border
                    # Apply row background
                    cell.fill = first_col_fill if ci == 1 else (even_fill if (row_idx - 1) % 2 == 0 else odd_fill)

            # Adjust column widths
            for ci, col in enumerate(headers, 1):
                max_length = max(
                    len(str(col)),
                    max((len(str(row.get(col, ''))) for row in export_rows), default=10)
                )
                adjusted_width = max_length * 1.2
                ws.column_dimensions[get_column_letter(ci)].width = adjusted_width

            # Save the workbook
            wb.save(file_path)
            self.logger.info(f"Pivot table exported to {file_path}")
            self.pivot_tab.status_label.setText(f"Exported to {file_path}")
            QMessageBox.information(self.pivot_tab, "Success", "Pivot table exported successfully!")
            
            # Ask to open the file
            if QMessageBox.question(self.pivot_tab, "Open File", "Open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                try:
                    if os.name == 'nt':  # Windows
                        os.startfile(file_path)
                    elif os.name == 'posix':  # macOS or Linux
                        os.system(f"open '{file_path}'" if platform.system() == "Darwin" else f"xdg-open '{file_path}'")
                except Exception as e:
                    self.logger.error(f"Failed to open file: {str(e)}")
                    QMessageBox.warning(self.pivot_tab, "Error", f"Failed to open file: {str(e)}")

        except Exception as e:
            self.logger.error(f"Failed to export pivot table: {str(e)}")
            self.pivot_tab.status_label.setText(f"Error: {str(e)}")
            QMessageBox.warning(self.pivot_tab, "Error", f"Failed to export pivot table: {str(e)}")