from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTableView, QAbstractItemView,
    QHeaderView, QScrollBar, QComboBox, QLineEdit, QDialog, QFileDialog, QMessageBox, QGroupBox, QProgressBar, QProgressDialog,
    QTabWidget, QScrollArea, QCheckBox
)
from PyQt6.QtCore import Qt, QAbstractTableModel, QTimer, QThread, pyqtSignal,pyqtSlot
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QFont, QColor
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import os
import platform
import math
from functools import reduce
import re
import logging

from .changeReport import ChangesReportDialog
from .column_filter import ColumnFilterDialog, FilterDialog

# Setup logging
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Global stylesheet
global_style = """
    QWidget {
        background-color: #F5F7FA;
        font-family: 'Inter', 'Segoe UI', sans-serif;
        font-size: 13px;
    }
    QGroupBox {
        font-weight: bold;
        color: #1A3C34;
        margin-top: 15px;
        border: 1px solid #D0D7DE;
        border-radius: 6px;
        padding: 10px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 0 5px;
        left: 10px;
    }
    QPushButton {
        background-color: #2E7D32;
        color: white;
        border: none;
        padding: 8px 16px;
        font-weight: 600;
        font-size: 13px;
        border-radius: 6px;
    }
    QPushButton:hover {
        background-color: #1B5E20;
    }
    QPushButton:disabled {
        background-color: #E0E0E0;
        color: #6B7280;
    }
    QTableView {
        background-color: #FFFFFF;
        border: 1px solid #D0D7DE;
        gridline-color: #E5E7EB;
        font-size: 12px;
        selection-background-color: #DBEAFE;
        selection-color: #1A3C34;
    }
    QHeaderView::section {
        background-color: #F9FAFB;
        font-weight: 600;
        color: #1A3C34;
        border: 1px solid #D0D7DE;
        padding: 6px;
    }
    QTableView::item:selected {
        background-color: #DBEAFE;
        color: #1A3C34;
    }
    QTableView::item {
        padding: 0px;
    }
    QLineEdit {
        padding: 6px;
        border: 1px solid #D0D7DE;
        border-radius: 4px;
        font-size: 13px;
    }
    QLineEdit:focus {
        border: 1px solid #2E7D32;
    }
    QLabel {
        font-size: 13px;
        color: #1A3C34;
    }
    QComboBox {
        padding: 6px;
        border: 1px solid #D0D7DE;
        border-radius: 4px;
        font-size: 13px;
    }
    QComboBox:focus {
        border: 1px solid #2E7D32;
    }
    QProgressBar {
        border: 1px solid #D0D7DE;
        border-radius: 4px;
        text-align: center;
    }
    QProgressBar::chunk {
        background-color: #2E7D32;
    }
"""

class FreezeTableWidget(QTableView):
    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.frozenTableView = QTableView(self)
        self.frozen_columns = 2  # NEW: Default to freeze two columns (checkbox + Solution Label)
        self.setModel(model)
        self.frozenTableView.setModel(model)
        self._is_dialog_open = False  # Flag to prevent multiple dialogs
        self.init()

        self.horizontalHeader().sectionResized.connect(self.updateSectionWidth)
        self.verticalHeader().sectionResized.connect(self.updateSectionHeight)
        self.frozenTableView.verticalScrollBar().valueChanged.connect(self.frozenVerticalScroll)
        self.verticalScrollBar().valueChanged.connect(self.mainVerticalScroll)

    def init(self):
        self.frozenTableView.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.frozenTableView.verticalHeader().hide()
        self.frozenTableView.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.viewport().stackUnder(self.frozenTableView)
        self.frozenTableView.setStyleSheet(global_style)
        self.frozenTableView.setSelectionModel(self.selectionModel())
        self.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.frozenTableView.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.frozenTableView.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.update_frozen_columns()
        self.frozenTableView.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozenTableView.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.frozenTableView.horizontalHeader().sectionClicked.connect(self.on_frozen_header_clicked)
        self.updateFrozenTableGeometry()
        self.frozenTableView.show()

    def update_frozen_columns(self):
        if self.model() is None or self.model().columnCount() < self.frozen_columns:
            self.frozenTableView.hide()
            logger.debug("Hiding frozen table due to insufficient columns")
            return
        for col in range(self.model().columnCount()):
            self.frozenTableView.setColumnHidden(col, col >= self.frozen_columns)
        for col in range(self.frozen_columns):
            column_width = self.columnWidth(col) if self.model().columnCount() > col else 100
            self.frozenTableView.setColumnWidth(col, column_width)
        self.frozenTableView.show()
        self.updateFrozenTableGeometry()
        logger.debug(f"Updated frozen columns: {self.frozen_columns}")

    def updateSectionWidth(self, logicalIndex, oldSize, newSize):
        if logicalIndex < self.frozen_columns:
            self.frozenTableView.setColumnWidth(logicalIndex, newSize)
            self.updateFrozenTableGeometry()
            self.frozenTableView.viewport().update()
        logger.debug(f"Section width updated for index {logicalIndex}: {newSize}")

    def updateSectionHeight(self, logicalIndex, oldSize, newSize):
        self.frozenTableView.setRowHeight(logicalIndex, newSize)

    def frozenVerticalScroll(self, value):
        self.viewport().stackUnder(self.frozenTableView)
        self.verticalScrollBar().setValue(value)
        self.frozenTableView.viewport().update()
        self.viewport().update()

    def mainVerticalScroll(self, value):
        self.viewport().stackUnder(self.frozenTableView)
        self.frozenTableView.verticalScrollBar().setValue(value)
        self.frozenTableView.viewport().update()
        self.viewport().update()

    def updateFrozenTableGeometry(self):
        if self.model() is None or self.model().columnCount() < self.frozen_columns:
            return
        total_width = sum(self.columnWidth(col) for col in range(self.frozen_columns))
        self.frozenTableView.setGeometry(
            self.verticalHeader().width() + self.frameWidth(),
            self.frameWidth(),
            total_width,
            self.viewport().height() + self.horizontalHeader().height()
        )
        self.frozenTableView.setFixedWidth(total_width)
        self.frozenTableView.viewport().update()
        logger.debug(f"Updated frozen table geometry with width: {total_width}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.updateFrozenTableGeometry()
        self.frozenTableView.viewport().update()

    def moveCursor(self, cursorAction, modifiers):
        current = super().moveCursor(cursorAction, modifiers)
        if cursorAction == QAbstractItemView.CursorAction.MoveLeft and current.column() >= self.frozen_columns:
            visual_x = self.visualRect(current).topLeft().x()
            frozen_width = sum(self.columnWidth(col) for col in range(self.frozen_columns))
            if visual_x < frozen_width:
                new_value = self.horizontalScrollBar().value() + visual_x - frozen_width
                self.horizontalScrollBar().setValue(int(new_value))
        return current

    def scrollTo(self, index, hint=QAbstractItemView.ScrollHint.EnsureVisible):
        if index.column() >= self.frozen_columns:
            super().scrollTo(index, hint)
        self.frozenTableView.viewport().update()

    def on_frozen_header_clicked(self, section):
        """Redirect frozen header click to ResultsFrame's header click handler"""
        logger.debug(f"Frozen header clicked for section: {section}")
        if self.model() is None:
            logger.warning("No model set for frozen table")
            QMessageBox.warning(self, "Error", "Table model not initialized.")
            return
        parent = self.parent().parent() if self.parent() else None
        logger.debug(f"ResultsFrame parent: {parent}, type: {type(parent).__name__ if parent else 'None'}")
        if section < self.frozen_columns and not self._is_dialog_open:
            if parent is not None and hasattr(parent, 'on_header_clicked'):
                self._is_dialog_open = True
                col_name = "Select" if section == 0 else "Solution Label"
                logger.debug(f"Calling ResultsFrame on_header_clicked for {col_name}")
                try:
                    parent.on_header_clicked(section, col_name=col_name)
                finally:
                    self._is_dialog_open = False
            else:
                logger.warning(f"Cannot call on_header_clicked. Parent: {parent}, has_method: {hasattr(parent, 'on_header_clicked') if parent else False}")
                QMessageBox.warning(self, "Error", "Cannot open filter dialog: ResultsFrame not found.")
        else:
            if section >= self.frozen_columns:
                logger.warning(f"Unexpected section {section} clicked in frozen table")
            if self._is_dialog_open:
                logger.debug("Dialog already open, ignoring click")

    def setModel(self, model):
        super().setModel(model)
        if self.frozenTableView is not None:
            self.frozenTableView.setModel(model)
            self.update_frozen_columns()
            self.updateFrozenTableGeometry()

class PandasModel(QAbstractTableModel):
    """Custom model to display pandas DataFrame in QTableView"""
    def __init__(self, data=pd.DataFrame(), format_value=None):
        super().__init__()
        self._data = data
        self._format_value = format_value
        self._checkboxes = [False] * len(data)  # List to track checkbox states

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data.columns) + 1  # +1 for checkbox column

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        col = index.column()
        row = index.row()
        if col == 0:  # Checkbox column
            if role == Qt.ItemDataRole.CheckStateRole:
                return Qt.CheckState.Checked if self._checkboxes[row] else Qt.CheckState.Unchecked
            return None
        else:  # Data columns
            value = self._data.iloc[row, col - 1]  # Shift by 1 for checkbox
            if role == Qt.ItemDataRole.DisplayRole:
                if self._format_value is not None:
                    return self._format_value(value)
                return str(value)
            elif role == Qt.ItemDataRole.BackgroundRole:
                if self._checkboxes[row]:  # Highlight checked rows
                    return QColor("#E6F3FA")  # Light blue for checked rows
                return QColor("#F9FAFB") if row % 2 else Qt.GlobalColor.white
        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if not index.isValid():
            return False

        row = index.row()
        col = index.column()

        # 1. چک‌باکس (ستون 0)
        if col == 0 and role == Qt.ItemDataRole.CheckStateRole:
            self._checkboxes[row] = (value == Qt.CheckState.Checked.value)
            self.dataChanged.emit(index, index)
            # به‌روزرسانی رنگ ردیف
            row_start = self.index(row, 0)
            row_end = self.index(row, self.columnCount() - 1)
            self.dataChanged.emit(row_start, row_end)
            return True

        # 2. ویرایش مقادیر عددی (ستون‌های > 0)
        if role == Qt.ItemDataRole.EditRole and col > 0:
            try:
                # تبدیل مقدار ورودی به float
                numeric_value = float(value)
                # ستون واقعی در DataFrame (چون ستون 0 = چک‌باکس)
                df_col_index = col - 1
                if df_col_index >= len(self._data.columns):
                    logger.warning(f"Column index {df_col_index} out of range for DataFrame with {len(self._data.columns)} columns.")
                    return False

                # به‌روزرسانی DataFrame
                self._data.iloc[row, df_col_index] = numeric_value

                # اطلاع‌رسانی به Qt
                self.dataChanged.emit(index, index)
                logger.debug(f"Updated PandasModel at row {row}, col {col} (df_col {df_col_index}) to {numeric_value}")
                return True

            except (ValueError, IndexError) as e:
                logger.warning(f"Failed to set data at row {row}, col {col}: {str(e)}")
                return False

        return False

    def flags(self, index):
        flags = super().flags(index)
        if index.column() == 0:
            flags |= Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled
        return flags

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                if section == 0:
                    return "Select"  # Checkbox header
                return str(self._data.columns[section - 1]) if section - 1 < len(self._data.columns) else ""
            return str(self._data.index[section]) if section < len(self._data.index) else ""
        return None

class DataWorker(QThread):
    data_ready = pyqtSignal(pd.DataFrame)
    error_occurred = pyqtSignal(str)

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        try:
            df = self.parent.compute_filtered_data()
            if df is None:
                logger.warning("compute_filtered_data returned None, emitting empty DataFrame")
                self.data_ready.emit(pd.DataFrame())
            else:
                self.data_ready.emit(df)
        except Exception as e:
            logger.error(f"Error in compute_filtered_data: {str(e)}")
            self.error_occurred.emit(str(e))

class ResultsFrame(QWidget):
    def __init__(self, app, parent=None):
        super().__init__(parent)
        self.app = app
        self.setStyleSheet(global_style)
        self.search_var = ""
        self.filter_field = "Solution Label"
        self.filter_values = {}
        self.column_filters = {}
        self.results_df = None  # DataFrame نتایج
        self.column_widths = {}
        self.column_backups = {}
        self.last_filtered_data = None
        self.last_pivot_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.decimal_places = "1"
        self.data_hash = None
        self.worker = None
        self.instance_id = id(self)
        logger.debug(f"ResultsFrame initialized with instance_id: {self.instance_id}")
        self.setup_ui()
        self.app.notify_data_changed = self.on_data_changed
        df = self.app.get_data()
        logger.debug(f"Initial data from app.get_data(): {df.shape if df is not None else 'None'}")
        self.show_processed_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        controls_group = QGroupBox("Table Controls")
        controls_layout = QHBoxLayout(controls_group)
        controls_layout.setSpacing(12)
        controls_layout.setContentsMargins(10, 10, 10, 10)

        search_label = QLabel("Search:")
        search_label.setFont(QFont("Segoe UI", 12))
        controls_layout.addWidget(search_label)
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Enter search term...")
        self.search_entry.setToolTip("Enter text to search in the pivot table")
        self.search_entry.setFixedWidth(200)
        self.search_entry.textChanged.connect(self.debounce_search)
        controls_layout.addWidget(self.search_entry)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(200)
        self.progress_bar.setVisible(False)
        controls_layout.addWidget(self.progress_bar)

        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.show_processed_data)

        filter_button = QPushButton("📌 Filter")
        filter_button.setToolTip("Filter the pivot table by Solution Label or Element")
        filter_button.setMinimumWidth(120)
        filter_button.clicked.connect(self.open_filter_window)
        controls_layout.addWidget(filter_button)

        clear_col_filters_btn = QPushButton("Clear Col Filters")
        clear_col_filters_btn.setToolTip("Clear all column filters")
        clear_col_filters_btn.setMinimumWidth(120)
        clear_col_filters_btn.clicked.connect(self.clear_column_filters)
        controls_layout.addWidget(clear_col_filters_btn)

        self.save_button = QPushButton("💾 Save Excel")
        self.save_button.setToolTip("Save the pivot table to an Excel file")
        self.save_button.clicked.connect(self.save_processed_excel)
        self.save_button.setMinimumWidth(120)
        controls_layout.addWidget(self.save_button)

        self.save_raw_button = QPushButton("💾 Save Raw Excel")
        self.save_raw_button.setToolTip("Save the raw table without pivoting")
        self.save_raw_button.clicked.connect(self.save_raw_excel)
        self.save_raw_button.setMinimumWidth(150)
        controls_layout.addWidget(self.save_raw_button)

        self.decimal_combo = QComboBox()
        self.decimal_combo.addItems(["0", "1", "2", "3"])
        self.decimal_combo.setCurrentText(self.decimal_places)
        self.decimal_combo.setFixedWidth(60)
        self.decimal_combo.setToolTip("Set the number of decimal places for numeric values")
        self.decimal_combo.currentTextChanged.connect(self.show_processed_data)

        report_button = QPushButton("Report Changes")
        report_button.clicked.connect(self.show_changes_report)
        controls_layout.addWidget(report_button)

        compare_button = QPushButton("Compare Rows")
        compare_button.setToolTip("Compare two selected rows and compute differences")
        compare_button.clicked.connect(self.compare_selected_rows)
        controls_layout.addWidget(compare_button)

        similar_button = QPushButton("Find Similar")
        similar_button.setToolTip("Find rows similar to the selected row")
        similar_button.clicked.connect(self.find_similar_rows)
        controls_layout.addWidget(similar_button)

        oreas_button = QPushButton("Compare with OREAS")
        oreas_button.setToolTip("Compare selected rows with OREAS in Find Similarity tab")
        oreas_button.clicked.connect(self.compare_with_oreas)
        controls_layout.addWidget(oreas_button)

        controls_layout.addWidget(self.decimal_combo)
        controls_layout.addStretch()

        layout.addWidget(controls_group)

        table_group = QGroupBox("Pivot Table")
        table_layout = QVBoxLayout(table_group)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.processed_table = FreezeTableWidget(PandasModel(), parent=self)
        self.processed_table.setStyleSheet(global_style)
        self.processed_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.processed_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.processed_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.processed_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.processed_table.setToolTip("Processed pivot table with filtered data")
        self.processed_table.setEnabled(False)
        self.processed_table.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
        self.processed_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        table_layout.addWidget(self.processed_table)

        layout.addWidget(table_group, stretch=1)
        self.setLayout(layout)

    def get_selected_checkbox_rows(self):
        model = self.processed_table.model()
        if model is None:
            return []
        selected_rows = [i for i in range(model.rowCount()) if model._checkboxes[i]]
        return selected_rows

    def compare_with_oreas(self):
        selected_rows = self.get_selected_checkbox_rows()
        if len(selected_rows) < 1:
            QMessageBox.warning(self, "Warning", "Please select at least one row using checkboxes to compare with OREAS.")
            return

        df = self.last_filtered_data
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data available for comparison.")
            return

        # ساختن DataFrame کنترل با تمام ستون‌ها (شامل طول موج‌ها)
        control_data = {'SAMPLE ID': []}
        for row_idx in selected_rows:
            selected_row = df.iloc[row_idx]
            solution_label = selected_row['Solution Label']
            control_data['SAMPLE ID'].append(solution_label)
            for col in df.columns:
                if col != 'Solution Label':
                    if col not in control_data:
                        control_data[col] = []
                    control_data[col].append(selected_row[col])

        control_df = pd.DataFrame(control_data)

        # ارسال به CompareTab
        self.app.compare_tab.set_control_from_results(control_df)

        # تغییر تب
        if hasattr(self.app, 'main_content'):
            self.app.main_content.switch_tab("Find similarity")
        else:
            QMessageBox.warning(self, "Error", "Cannot switch to Find Similarity tab.")

        QMessageBox.information(self, "Success", f"Selected {len(selected_rows)} row(s) sent to Find Similarity tab as control. Comparing with OREAS.")

    def show_changes_report(self):
        dialog = ChangesReportDialog(self.app, self)
        dialog.exec()

    def update_table_display(self):
        pass

    def connect_to_crm_check(self, crm_check):
        logger.debug(f"Connecting CrmCheck to ResultsFrame instance_id: {self.instance_id}")
        crm_check.data_changed.connect(self.on_crm_data_changed)

    def on_crm_data_changed(self):
        logger.debug(f"Data changed in CrmCheck, updating ResultsFrame table instance_id: {self.instance_id}")
        self.update_table(self.last_filtered_data)

    def debounce_search(self, text):
        self.search_var = text
        self.search_timer.start(500)

    def format_value(self, x):
        try:
            value = float(x)
            decimal_places = int(self.decimal_combo.currentText())
            formatted = f"{value:.{decimal_places}f}".rstrip('0').rstrip('.')
            return formatted
        except (ValueError, TypeError):
            return str(x)

    def is_numeric(self, value):
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False

    def reset_filter_cache(self):
        self.last_filtered_data = None
        self._last_cache_key = None
        self.data_hash = None
        self.column_filters.clear()
        self.filter_values.clear()
        logger.debug(f"Reset filter cache for instance_id: {self.instance_id}")

    def data_changed(self):
        logger.debug(f"ResultsFrame data_changed for instance_id: {self.instance_id}")
        self.reset_filter_cache()
        self.last_pivot_data = None
        self.show_processed_data()

    def on_data_changed(self):
        logger.debug(f"on_data_changed triggered for instance_id: {self.instance_id}")
        self.reset_filter_cache()
        self.last_pivot_data = None
        self.show_processed_data()

    def compute_filtered_data(self):
        logger.debug(f"Starting compute_filtered_data for instance_id: {self.instance_id}")
        
        df = self.app.get_data()
        if df is None or df.empty:
            logger.warning("No data available from app.get_data()")
            self.last_pivot_data = None
            return pd.DataFrame()

        required_columns = ['Solution Label', 'Element', 'Corr Con']
        if not all(col in df.columns for col in required_columns):
            logger.error(f"DataFrame missing required columns: {required_columns}")
            self.last_pivot_data = None
            return pd.DataFrame()

        hash_columns = required_columns + ['row_id', 'original_index'] if 'row_id' in df.columns else required_columns
        new_hash = str(pd.util.hash_pandas_object(df[hash_columns]).sum())
        logger.debug(f"Computed data hash: {new_hash}")

        logger.debug(f"Current column_filters: {self.column_filters}")
        logger.debug(f"Current filter_values: {self.filter_values}")

        if new_hash != self.data_hash or self.last_pivot_data is None:
            logger.debug("Data changed or no pivot data, recomputing pivot")
            df_filtered = df[df['Type'].isin(['Samp', 'Sample'])].copy()
            logger.debug(f"After Type filter, df_filtered shape: {df_filtered.shape}")
            df_filtered = df_filtered[
                (~df_filtered['Solution Label'].isin(self.app.get_excluded_samples())) &
                (~df_filtered['Solution Label'].isin(self.app.get_excluded_volumes())) &
                (~df_filtered['Solution Label'].isin(self.app.get_excluded_dfs()))
            ]
            logger.debug(f"After exclusion filters, df_filtered shape: {df_filtered.shape}")

            if df_filtered.empty:
                logger.warning("No data after initial filtering")
                self.last_pivot_data = None
                return pd.DataFrame()

            if 'original_index' not in df_filtered.columns:
                df_filtered['original_index'] = df_filtered.index
            df_filtered['Element'] = df_filtered['Element'].str.split('_').str[0]
            df_filtered = df_filtered.reset_index(drop=True)

            most_common_sizes = {}
            for solution_label in df_filtered['Solution Label'].unique():
                df_subset = df_filtered[df_filtered['Solution Label'] == solution_label]
                counts = df_subset['Element'].value_counts().values
                total_rows = len(df_subset)
                g = reduce(math.gcd, counts) if len(counts) > 0 else 1
                most_common_sizes[solution_label] = total_rows // g if g > 0 and total_rows % g == 0 else total_rows

            df_filtered['set_size'] = df_filtered['Solution Label'].map(most_common_sizes)

            group_counts = df_filtered.groupby(['Solution Label', df_filtered.groupby('Solution Label').cumcount() // df_filtered['set_size'], 'Element']).size()
            has_repeats = (group_counts > 1).any()
            logger.debug(f"Has repeated elements: {has_repeats}")

            def clean_label(label):
                m = re.search(r'(\d+)', str(label).replace(' ', ''))
                return f"{label.split()[0]} {m.group(1)}" if m else label

            if self.solution_label_order is None or not self.solution_label_order:
                self.solution_label_order = sorted(df_filtered['Solution Label'].drop_duplicates().apply(clean_label).tolist())

            value_column = 'Corr Con'
            if value_column not in df_filtered.columns:
                logger.error(f"Column '{value_column}' not found in data")
                self.last_pivot_data = None
                return pd.DataFrame()

            if not has_repeats:
                df_filtered['unique_id'] = df_filtered.groupby(['Solution Label', 'Element']).cumcount()
                if self.element_order is None or not self.element_order:
                    self.element_order = df_filtered['Element'].drop_duplicates().tolist()

                pivot_data = df_filtered.pivot_table(
                    index=['Solution Label', 'unique_id'],
                    columns='Element',
                    values=value_column,
                    aggfunc='first',
                    sort=False
                ).reset_index()

                pivot_data = pivot_data.merge(
                    df_filtered[['Solution Label', 'unique_id', 'original_index']].drop_duplicates(),
                    on=['Solution Label', 'unique_id'],
                    how='left'
                ).sort_values('original_index').drop(columns=['unique_id', 'original_index'])

            else:
                df_filtered['group_id'] = df_filtered.groupby('Solution Label').cumcount() // df_filtered['set_size']

                element_counts = df_filtered.groupby(['Solution Label', 'group_id', 'Element']).size().reset_index(name='count')
                df_filtered = df_filtered.merge(
                    element_counts[['Solution Label', 'group_id', 'Element', 'count']],
                    on=['Solution Label', 'group_id', 'Element'],
                    how='left'
                )
                df_filtered['count'] = df_filtered['count'].fillna(1).astype(int)
                df_filtered['element_count'] = df_filtered.groupby(['Solution Label', 'group_id', 'Element']).cumcount() + 1
                df_filtered['Element_with_id'] = df_filtered.apply(
                    lambda x: f"{x['Element']}_{x['element_count']}" if x['count'] > 1 else x['Element'],
                    axis=1
                )

                expected_columns_dict = {}
                for solution_label in df_filtered['Solution Label'].unique():
                    expected_size = most_common_sizes.get(solution_label, 1)
                    set_sizes_subset = df_filtered[df_filtered['Solution Label'] == solution_label].groupby('group_id').size().reset_index(name='set_size')
                    valid_groups = set_sizes_subset[set_sizes_subset['set_size'] == expected_size]['group_id']
                    if not valid_groups.empty:
                        first_group_id = valid_groups.min()
                        first_set_elements = df_filtered[
                            (df_filtered['Solution Label'] == solution_label) & 
                            (df_filtered['group_id'] == first_group_id)
                        ]['Element_with_id'].unique().tolist()
                        expected_columns_dict[solution_label] = first_set_elements
                    else:
                        expected_columns_dict[solution_label] = []

                if self.element_order is None or not self.element_order:
                    self.element_order = list(set().union(*[set(cols) for cols in expected_columns_dict.values()]))

                pivot_dfs = []
                min_index_per_group = {}
                for solution_label, expected_columns in expected_columns_dict.items():
                    if not expected_columns:
                        logger.debug(f"No valid columns for Solution Label: {solution_label}")
                        continue
                    df_subset = df_filtered[df_filtered['Solution Label'] == solution_label].copy()
                    min_index_per_group[solution_label] = df_subset.groupby('group_id')['original_index'].min().to_dict()
                    pivot_subset = df_subset.pivot_table(
                        index=['Solution Label', 'group_id'],
                        columns='Element_with_id',
                        values=value_column,
                        aggfunc='first',
                        sort=False
                    ).reset_index()
                    pivot_subset = pivot_subset.reindex(columns=['Solution Label', 'group_id'] + expected_columns)
                    pivot_subset['min_original_index'] = pivot_subset['group_id'].map(min_index_per_group[solution_label])
                    pivot_dfs.append(pivot_subset)

                if not pivot_dfs:
                    logger.error("No valid pivot tables created")
                    self.last_pivot_data = None
                    return pd.DataFrame()
                
                pivot_data = pd.concat(pivot_dfs, ignore_index=True)
                if 'min_original_index' in pivot_data.columns:
                    pivot_data = pivot_data.sort_values(by='min_original_index').reset_index(drop=True)
                columns_to_drop = [col for col in ['group_id', 'min_original_index'] if col in pivot_data.columns]
                if columns_to_drop:
                    pivot_data = pivot_data.drop(columns=columns_to_drop)

                self.last_pivot_data = pivot_data
                self.data_hash = new_hash
                self.last_filtered_data = None
                self._last_cache_key = None
                logger.debug(f"Pivot data shape: {pivot_data.shape}")
        else:
            pivot_data = self.last_pivot_data
            logger.debug("Using cached pivot data")

        if pivot_data is None or pivot_data.empty:
            logger.warning("Pivot data is None or empty")
            return pd.DataFrame()

        search_text = self.search_var.lower().strip()
        filter_field = self.filter_field
        selected_values = [k for k, v in self.filter_values.get(filter_field, {}).items() if v]
        logger.debug(f"Search text: {search_text}, Filter field: {filter_field}, Selected values: {selected_values}")

        cache_key = (
            search_text,
            filter_field,
            tuple(sorted(selected_values)),
            self.data_hash,
            str(self.column_filters)
        )
        if cache_key == self._last_cache_key and self.last_filtered_data is not None:
            logger.debug("Returning cached filtered data")
            return self.last_filtered_data

        filtered_pivot = pivot_data.copy()

        for col_name, col_filter in self.column_filters.items():
            if col_name in filtered_pivot.columns:
                col_data = pd.to_numeric(filtered_pivot[col_name], errors='coerce')
                is_numeric_col = col_data.notna().any() and col_name != 'Solution Label'
                logger.debug(f"Applying filter for column {col_name}, is_numeric: {is_numeric_col}")

                if 'min_val' in col_filter and col_filter['min_val'] is not None and is_numeric_col:
                    min_mask = (col_data >= col_filter['min_val']) | col_data.isna()
                    filtered_pivot = filtered_pivot[min_mask]
                    logger.debug(f"Applied min filter {col_filter['min_val']} on column {col_name}")

                if 'max_val' in col_filter and col_filter['max_val'] is not None and is_numeric_col:
                    max_mask = (col_data <= col_filter['max_val']) | col_data.isna()
                    filtered_pivot = filtered_pivot[max_mask]
                    logger.debug(f"Applied max filter {col_filter['max_val']} on column {col_name}")

                if 'selected_values' in col_filter and col_filter['selected_values']:
                    if is_numeric_col:
                        selected_values_set = {float(val) for val in col_filter['selected_values'] if self.is_numeric(str(val))}
                        selected_mask = col_data.isin(selected_values_set)
                    else:
                        selected_values_set = set(col_filter['selected_values'])
                        selected_mask = filtered_pivot[col_name].isin(selected_values_set)
                    filtered_pivot = filtered_pivot[selected_mask]
                    logger.debug(f"Applied selected values filter on column {col_name}: {selected_values_set}")

        logger.debug(f"After column filtering - filtered_pivot shape: {filtered_pivot.shape}")

        if search_text:
            search_columns = ['Solution Label'] + [col for col in filtered_pivot.columns if col != 'Solution Label']
            mask = filtered_pivot[search_columns].apply(
                lambda col: col.astype(str).str.lower().str.contains(search_text, na=False)
            ).any(axis=1)
            filtered_pivot = filtered_pivot[mask]
            logger.debug(f"After search filtering - filtered_pivot shape: {filtered_pivot.shape}")

        if filter_field and selected_values:
            if filter_field == 'Solution Label':
                selected_order = [x for x in self.solution_label_order if x in selected_values and x in filtered_pivot['Solution Label'].values]
                filtered_pivot = filtered_pivot[filtered_pivot['Solution Label'].isin(selected_values)]
                if not filtered_pivot.empty:
                    filtered_pivot['Solution Label'] = pd.Categorical(
                        filtered_pivot['Solution Label'],
                        categories=selected_order,
                        ordered=True
                    )
                    filtered_pivot = filtered_pivot.sort_values('Solution Label').reset_index(drop=True)
                logger.debug(f"After Solution Label filtering - filtered_pivot shape: {filtered_pivot.shape}")
            elif filter_field == 'Element':
                columns_to_keep = ['Solution Label'] + [col for col in self.element_order if col in selected_values and col in filtered_pivot.columns]
                filtered_pivot = filtered_pivot[columns_to_keep]
                logger.debug(f"After Element filtering - filtered_pivot shape: {filtered_pivot.shape}")

        filtered_pivot = filtered_pivot.drop_duplicates().reset_index(drop=True)
        self.last_filtered_data = filtered_pivot
        self._last_cache_key = cache_key
        logger.debug(f"Final filtered_pivot shape: {filtered_pivot.shape}")

        return filtered_pivot

    @pyqtSlot(dict)
    def update_results_from_compare(self, updates):
        """
        updates = {
            "Control123": {"Al 368.098": 122.4, "Fe 238.204": 83.3}
        }
        """
        print("omid1 :", updates)

        if not hasattr(self, 'processed_table') or not self.processed_table:
            logger.warning("processed_table not initialized.")
            return

        model = self.processed_table.model()
        if not model:
            logger.warning("processed_table model is None.")
            return

        if not isinstance(model, PandasModel):
            logger.error("Expected PandasModel, but got different model type.")
            return

        updated = 0

        for control_id, col_updates in updates.items():
            # ستون SAMPLE ID = ستون 1 (چون ستون 0 = چک‌باکس)
            sample_id_col = 1
            found_row = -1

            for row in range(model.rowCount()):
                index = model.index(row, sample_id_col)
                current_id = model.data(index, Qt.ItemDataRole.DisplayRole)
                if current_id == control_id:
                    found_row = row
                    break

            if found_row == -1:
                logger.warning(f"Control ID {control_id} not found in Results table.")
                continue

            # به‌روزرسانی هر ستون
            for col_name, new_val in col_updates.items():
                col_idx = -1
                # پیدا کردن ستون با نام
                for c in range(model.columnCount()):
                    header_index = model.createIndex(0, c)  # یا model.index(0, c)
                    header_text = model.headerData(c, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                    if header_text == col_name:
                        col_idx = c
                        break

                if col_idx == -1:
                    logger.warning(f"Column {col_name} not found in Results table.")
                    continue

                # به‌روزرسانی مقدار در PandasModel
                index = model.index(found_row, col_idx)
                success = model.setData(index, f"{new_val:.2f}", Qt.ItemDataRole.EditRole)
                if success:
                    updated += 1
                else:
                    logger.warning(f"Failed to set data at row {found_row}, col {col_idx}")

        if updated > 0:
            logger.debug(f"Results table updated: {updated} cell(s) changed via PandasModel.")
            # اختیاری: به‌روزرسانی UI
            self.processed_table.viewport().update()
            # QMessageBox.information(self, "Updated", f"{updated} values updated.")
        else:
            logger.info("No updates applied to Results table.")

    def show_processed_data(self):
        if self.worker is not None and self.worker.isRunning():
            logger.debug(f"Worker already running for instance_id: {self.instance_id}, skipping show_processed_data")
            return

        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        self.search_entry.setEnabled(False)
        self.worker = DataWorker(self)
        self.worker.data_ready.connect(self.update_table)
        self.worker.error_occurred.connect(self.show_error)
        self.worker.finished.connect(self.on_worker_finished)
        logger.debug(f"Starting DataWorker for instance_id: {self.instance_id}")
        self.worker.start()

    def on_worker_finished(self):
        self.progress_bar.setVisible(False)
        self.search_entry.setEnabled(True)
        if hasattr(self.app, 'notify_data_changed'):
            self.app.notify_data_changed()
        logger.debug(f"Worker finished for instance_id: {self.instance_id}")

    def update_table(self, df):
        logger.debug(f"Updating table for instance_id: {self.instance_id}, data shape: {df.shape if df is not None else 'None'}")
        if df is None or df.empty:
            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["Status"])
            model.appendRow([QStandardItem("No data available after filtering")])
            self.processed_table.setModel(model)
            self.processed_table.setColumnWidth(0, 150)
            self.column_widths = {"Status": 150}
            self.processed_table.frozenTableView.setModel(model)
            self.processed_table.update_frozen_columns()
            self.processed_table.setEnabled(False)
            logger.warning("Table updated with no data due to filtering")
            return

        columns = list(df.columns)
        self.column_widths = {}
        self.processed_table.setColumnWidth(0, 50)  # Checkbox column width
        self.column_widths["Select"] = 50
        self.processed_table.setColumnWidth(1, 150)  # Solution Label column width
        self.column_widths["Solution Label"] = 150

        for col_idx, col in enumerate(columns, 1):
            if col == 'Solution Label':
                continue
            sample_data = df[col].dropna().astype(str)
            max_width = max([len(str(col))] + [len(str(self.format_value(x))) for x in sample_data], default=10)
            pixel_width = min(max_width * 10, 300)
            self.column_widths[col] = pixel_width
            self.processed_table.setColumnWidth(col_idx, pixel_width)

        model = PandasModel(df, format_value=self.format_value)
        self.processed_table.setModel(model)
        self.processed_table.frozenTableView.setModel(model)
        self.processed_table.update_frozen_columns()
        
        self.processed_table.model().layoutChanged.emit()
        self.processed_table.frozenTableView.model().layoutChanged.emit()
        self.processed_table.viewport().update()
        self.processed_table.frozenTableView.viewport().update()
        self.processed_table.setEnabled(True)
        
        if hasattr(self.app, 'notify_data_changed'):
            self.app.notify_data_changed()

    def show_error(self, message):
        self.progress_bar.setVisible(False)
        self.search_entry.setEnabled(True)
        QMessageBox.warning(self, "Error", message)
        logger.error(f"Error shown for instance_id: {self.instance_id}: {message}")

    def open_filter_window(self):
        df = self.app.get_data()
        if df is None:
            QMessageBox.warning(self, "Warning", "No data to filter!")
            logger.warning(f"No data to filter in open_filter_window for instance_id: {self.instance_id}")
            return

        dialog = FilterDialog(
            self, self.filter_values, self.filter_field,
            self.solution_label_order, self.element_order
        )
        dialog.exec()
        self.filter_field = dialog.filter_combo.currentText()
        logger.debug(f"Filter window closed, filter_field set to {self.filter_field} for instance_id: {self.instance_id}")

    def on_header_clicked(self, section, col_name=None):
        logger.debug(f"on_header_clicked called with section: {section}, col_name: {col_name}, instance_id: {self.instance_id}")
        data_source = self.last_pivot_data if self.last_pivot_data is not None else self.last_filtered_data
        if data_source is None or data_source.empty:
            logger.warning("No pivot or filtered data available for filtering")
            QMessageBox.warning(self, "Warning", "No data available to filter!")
            return
        
        if col_name is None:
            model = self.processed_table.model()
            if model is not None:
                col_name = model.headerData(section, Qt.Orientation.Horizontal)
            if col_name is None or col_name == "":
                logger.warning(f"No valid column name for section {section}")
                QMessageBox.warning(self, "Warning", f"No valid column name for section {section}")
                return
        
        logger.debug(f"Opening ColumnFilterDialog for column: {col_name}")
        dialog = ColumnFilterDialog(self, col_name)
        dialog.exec()

    def clear_column_filters(self):
        self.column_filters.clear()
        self.show_processed_data()
        QMessageBox.information(self, "Filters Cleared", "All column filters have been cleared.")
        logger.debug(f"Column filters cleared for instance_id: {self.instance_id}")

    def compare_selected_rows(self):
        selected_rows = self.get_selected_checkbox_rows()
        if len(selected_rows) != 2:
            QMessageBox.warning(self, "Warning", "Please select exactly two rows using checkboxes to compare.")
            return

        df = self.last_filtered_data
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data available for comparison.")
            return

        row1 = df.iloc[selected_rows[0]]
        row2 = df.iloc[selected_rows[1]]
        label1 = row1['Solution Label']
        label2 = row2['Solution Label']

        diff_df = pd.DataFrame({
            'Column': df.columns,
            'Row1 Value': row1.values,
            'Row2 Value': row2.values,
            'Difference': np.nan
        })

        for col in df.columns:
            if col != 'Solution Label':
                df[col] = pd.to_numeric(df[col], errors='coerce')

        numeric_cols = df.columns[df.dtypes.apply(pd.api.types.is_numeric_dtype)]
        numeric_cols = [col for col in numeric_cols if col != 'Solution Label']

        for col in numeric_cols:
            try:
                val1 = float(row1[col]) if pd.notna(row1[col]) else np.nan
                val2 = float(row2[col]) if pd.notna(row2[col]) else np.nan
                if pd.notna(val1) and pd.notna(val2):
                    diff = val1 - val2
                    diff_df.loc[diff_df['Column'] == col, 'Difference'] = diff
            except (ValueError, TypeError):
                pass

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Comparison: {label1} vs {label2}")
        dialog_layout = QVBoxLayout(dialog)

        diff_table = QTableView()
        diff_model = PandasModel(diff_df, format_value=self.format_value)
        diff_table.setModel(diff_model)
        diff_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        dialog_layout.addWidget(diff_table)

        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        dialog_layout.addWidget(close_button)

        dialog.resize(800, 600)
        dialog.exec()

    def find_similar_rows(self):
        selected_rows = self.get_selected_checkbox_rows()
        if len(selected_rows) < 1:
            QMessageBox.warning(self, "Warning", "Please select at least one row using checkboxes to find similar rows.")
            return

        row_idx = selected_rows[0]
        df = self.last_filtered_data
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data available for similarity search.")
            return

        selected_row = df.iloc[row_idx]
        label = selected_row['Solution Label']

        data_numeric = df.copy()
        for col in data_numeric.columns:
            if col != 'Solution Label':
                data_numeric[col] = pd.to_numeric(data_numeric[col], errors='coerce')

        numeric_cols = data_numeric.columns[data_numeric.dtypes.apply(pd.api.types.is_numeric_dtype)]
        numeric_cols = [col for col in numeric_cols if col != 'Solution Label']

        if not numeric_cols:
            QMessageBox.warning(self, "Warning", "No numeric columns available for similarity computation.")
            return

        data_numeric = data_numeric[numeric_cols].fillna(0).astype(float)
        selected_vec = data_numeric.iloc[row_idx].values
        distances = np.linalg.norm(data_numeric.values - selected_vec, axis=1)
        df['Distance'] = distances
        similar_df = df.sort_values('Distance').reset_index(drop=True)
        similar_df = similar_df.drop(columns=['Distance'])

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Similar Rows to {label}")
        dialog_layout = QVBoxLayout(dialog)

        similar_table = QTableView()
        similar_model = PandasModel(similar_df, format_value=self.format_value)
        similar_table.setModel(similar_model)
        similar_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        dialog_layout.addWidget(similar_table)

        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        dialog_layout.addWidget(close_button)

        dialog.resize(800, 600)
        dialog.exec()

    def save_processed_excel(self):
        df = self.last_filtered_data
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data to save!")
            logger.warning(f"No data to save in save_processed_excel for instance_id: {self.instance_id}")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Processed Pivot Table"

                header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                first_column_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
                odd_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                header_font = Font(name="Segoe UI", size=12, bold=True)
                cell_font = Font(name="Segoe UI", size=12)
                cell_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )

                headers = list(df.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15

                for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                    fill = even_row_fill if (row_idx - 1) % 2 == 0 else odd_row_fill
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if pd.isna(value):
                            cell.value = None
                        else:
                            try:
                                float_value = float(value)
                                cell.value = float_value
                                decimal_places = int(self.decimal_combo.currentText())
                                cell.number_format = f"0.{'0' * decimal_places}"
                                cell.value = self.format_value(value)
                            except ValueError:
                                cell.value = str(value)
                        cell.font = cell_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        if col_idx == 1:
                            cell.fill = first_column_fill
                        else:
                            cell.fill = fill

                wb.save(file_path)
                QMessageBox.information(self, "Success", "Processed pivot table saved successfully!")
                logger.debug(f"Processed pivot table saved to {file_path} for instance_id: {self.instance_id}")

                if QMessageBox.question(self, "Open File", "Would you like to open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                    try:
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(file_path)
                        elif system == "Darwin":
                            os.system(f"open {file_path}")
                        else:
                            os.system(f"xdg-open {file_path}")
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")
                        logger.error(f"Failed to open file {file_path}: {str(e)}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save: {str(e)}")
                logger.error(f"Failed to save processed excel: {str(e)}")

    def save_raw_excel(self):
        df = self.app.get_data()
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No raw data to save!")
            logger.warning(f"No raw data to save for instance_id: {self.instance_id}")
            return

        df_filtered = df[df['Type'].isin(['Samp', 'Sample'])].copy()
        df_filtered = df_filtered[
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_samples())) &
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_volumes())) &
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_dfs()))
        ]

        if df_filtered.empty:
            QMessageBox.warning(self, "Warning", "No filtered raw data to save!")
            logger.warning(f"No filtered raw data to save for instance_id: {self.instance_id}")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Raw Excel File", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            try:
                df_filtered.to_excel(file_path, index=False)
                QMessageBox.information(self, "Success", "Raw table saved successfully!")
                logger.debug(f"Raw table saved to {file_path} for instance_id: {self.instance_id}")

                if QMessageBox.question(self, "Open File", "Would you like to open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                    try:
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(file_path)
                        elif system == "Darwin":
                            os.system(f"open {file_path}")
                        else:
                            os.system(f"xdg-open {file_path}")
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")
                        logger.error(f"Failed to open raw excel file {file_path}: {str(e)}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save raw data: {str(e)}")
                logger.error(f"Failed to save raw excel: {str(e)}")

    def reset_cache(self):
        self.last_filtered_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.column_widths = {}
        self.filter_values = {}
        self.column_filters = {}
        self.search_var = ""
        self.data_hash = None
        logger.debug(f"Reset cache for instance_id: {self.instance_id}, last_pivot_data preserved")

    def reset_state(self):
        logger.debug(f"Resetting ResultsFrame state for instance_id: {self.instance_id}")
        self.search_var = ""
        self.filter_field = "Solution Label"
        self.filter_values = {}
        self.column_filters = {}
        self.column_widths = {}
        self.column_backups = {}
        self.last_filtered_data = None
        self.last_pivot_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.decimal_places = "1"
        self.data_hash = None

        if hasattr(self, 'search_entry'):
            self.search_entry.setText("")
            logger.debug("Search entry cleared")
        if hasattr(self, 'decimal_combo'):
            self.decimal_combo.setCurrentText(self.decimal_places)
            logger.debug(f"Decimal combo reset to {self.decimal_places}")
        if hasattr(self, 'processed_table'):
            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["Status"])
            model.appendRow([QStandardItem("No data available after filtering")])
            self.processed_table.setModel(model)
            self.processed_table.setColumnWidth(0, 150)
            self.processed_table.frozenTableView.setModel(model)
            self.processed_table.update_frozen_columns()
            self.processed_table.setEnabled(False)
            logger.debug("Processed table reset with 'No data' message")

        if self.worker is not None and self.worker.isRunning():
            self.worker.terminate()
            self.worker = None
            self.progress_bar.setVisible(False)
            self.search_entry.setEnabled(True)
            logger.debug("Terminated running worker thread")

        if hasattr(self.app, 'notify_data_changed'):
            self.app.notify_data_changed()
            logger.debug("Notified data change")

        logger.debug(f"ResultsFrame state reset completed for instance_id: {self.instance_id}")